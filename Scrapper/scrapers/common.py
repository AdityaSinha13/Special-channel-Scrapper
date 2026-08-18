"""
Shared helpers used by every platform scraper module:
 - Chrome driver factory (stealth options, shared across Blinkit/Swiggy/Flipkart)
 - Excel/CSV writer with auto column width
 - Generic ID/URL loader from an uploaded file or pasted text
 - A tiny `StopFlag` so the Streamlit UI can interrupt a running scrape
"""
from __future__ import annotations

import re
from pathlib import Path
from io import BytesIO

import pandas as pd


class StopFlag:
    """Mutable flag the UI can flip to True to abort a running scrape loop."""
    def __init__(self):
        self.stop = False


def make_chrome_driver(headless: bool = False, window_size: str = "1440,900"):
    """Single stealth Chrome factory reused by Blinkit / Swiggy / Flipkart.
    (Amazon primarily uses `requests`, but can reuse this for its optional
    seller-scraping Selenium fallback.)
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager

    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(f"--window-size={window_size}")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--lang=en-IN")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )
    return driver


def clean_price(text) -> str:
    if not text:
        return ""
    cleaned = re.sub(r"[^\d.]", "", str(text).replace(",", ""))
    parts = cleaned.split(".")
    if len(parts) > 2:
        cleaned = parts[0] + "." + "".join(parts[1:])
    return cleaned


def load_ids_from_dataframe(df: pd.DataFrame, id_col_names=("pid", "pids", "asin", "fsn", "id")) -> list[str]:
    """Pick the most likely ID column out of an uploaded sheet, default to column 0."""
    col = next(
        (c for c in df.columns if str(c).strip().lower() in id_col_names),
        df.columns[0],
    )
    vals = []
    for v in df[col].dropna().astype(str):
        v = v.strip()
        if v and v.lower() != "nan":
            vals.append(v)
    return vals


def read_ids_file(uploaded_file) -> list[str]:
    """Accepts a Streamlit UploadedFile (csv/xlsx/txt) and returns a raw list of
    strings (could be bare IDs or full URLs — each platform module normalises
    these itself via its own `extract_id()` regexes)."""
    name = uploaded_file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file, dtype=str)
        return load_ids_from_dataframe(df)
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
        return load_ids_from_dataframe(df)
    # plain text, one ID/URL per line
    raw = uploaded_file.read().decode("utf-8", errors="ignore")
    return [line.strip() for line in raw.splitlines() if line.strip()]


def parse_pasted_ids(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data", header_color: str | None = None) -> bytes:
    """Return an in-memory .xlsx (auto column width, optional header fill,
    frozen header row) so Streamlit can offer it as a direct download."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        if header_color:
            fill = PatternFill("solid", fgColor=header_color)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
        ws.freeze_panes = "A2"
    return buf.getvalue()


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")
